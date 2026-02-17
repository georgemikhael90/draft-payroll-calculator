import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from utils import ServiceCategory
from PIL import Image as PILImage
import io

def add_header_logo(canvas, doc):
    """Add logo to the header of each page"""
    img = PILImage.open('attached_assets/NEW LOGO SMALL.png')
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)

    header_logo = Image(img_buffer)
    header_logo.drawHeight = 1.5*inch
    header_logo.drawWidth = 1.5*inch
    header_logo.drawOn(canvas, doc.leftMargin, doc.pagesize[1] - 2*inch)

def add_watermark(canvas, doc):
    """
    Add a 'Not for Official Use' watermark diagonally across the page.
    The watermark is semi-transparent to ensure it doesn't obstruct readability.
    """
    canvas.saveState()
    
    # Set watermark properties
    canvas.setFont('Helvetica-Bold', 36)  # Large bold font
    canvas.setFillColorRGB(0.85, 0.85, 0.85, alpha=0.25)  # Light gray with 25% opacity
    
    # Get page dimensions
    page_width, page_height = doc.pagesize
    
    # Position and rotate the watermark diagonally
    canvas.translate(page_width/2, page_height/2)  # Move to center of page
    canvas.rotate(45)  # Rotate 45 degrees
    
    # Draw the watermark text
    canvas.drawCentredString(0, 0, "NOT FOR OFFICIAL USE")
    
    canvas.restoreState()

def format_currency(amount):
    """Format amount as currency"""
    return "${:.2f}".format(float(amount))

def get_special_duty_text(hazardous_duty, hardship_duty, at_border, present_this_month):
    """Generate text describing special duty status"""
    if not present_this_month:
        return "No Special Duty Pay (Not Present)"

    duties = []
    if hazardous_duty:
        duties.append("✓ Hazardous Duty Pay ($1000.00/month)")
    if hardship_duty:
        duties.append("✓ Hardship Duty Pay ($500.00/month)")
    if at_border:
        duties.append("✓ Border Duty Pay ($225.00/month)")

    return "\n".join(duties) if duties else "No Special Duty Pay Selected"

def generate_pdf_report(pay_info, service_category, military_grade, years_of_service, start_date, end_date, 
                       has_dependents, hazardous_duty, hardship_duty, at_border,
                       sm_name, sm_dodid, sm_task_force, sm_company,
                       is_correction=False, original_details=None):
    """Generate a PDF report of the pay calculation"""
    print("Debug: Starting PDF report generation...")
    print(f"Debug: Is correction report: {is_correction}")
    print(f"Debug: Pay info keys: {pay_info.keys() if pay_info else 'No pay info'}")

    try:
        filename = f"pay_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)

        def onFirstPage(canvas, doc):
            add_header_logo(canvas, doc)
            add_watermark(canvas, doc)  # Add watermark to first page

        def onLaterPages(canvas, doc):
            add_header_logo(canvas, doc)
            add_watermark(canvas, doc)  # Add watermark to all pages

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        elements.append(Paragraph("SAD Pay Correction Report" if is_correction else "SAD Pay Calculator Report", title_style))
        elements.append(Spacer(1, 12))

        print("Debug: Basic report structure created...")

        # Service Member Information
        if any([sm_name, sm_dodid, sm_task_force, sm_company]):
            elements.append(Paragraph("Service Member Information", styles['Heading2']))
            sm_style = ParagraphStyle('ServiceMember', parent=styles['Normal'], fontSize=12, spaceAfter=6)

            sm_details = [
                f"Name: {sm_name}" if sm_name else None,
                f"DOD ID: {sm_dodid}" if sm_dodid else None,
                f"Task Force: {sm_task_force}" if sm_task_force else None,
                f"Company: {sm_company}" if sm_company else None
            ]

            for detail in sm_details:
                if detail:
                    elements.append(Paragraph(detail, sm_style))
            elements.append(Spacer(1, 12))

        print("Debug: Service member information added...")

        if is_correction:
            print("Debug: Processing correction report details...")
            # Add original pay details
            elements.append(Paragraph("Original Pay Details", styles['Heading2']))

            present = original_details.get('present_this_month', False)
            status_text = get_special_duty_text(
                original_details['hazardous_duty'] and present,
                original_details['hardship_duty'] and present,
                original_details['at_border'] and present,
                present
            )

            orig_info = [
                f"Service Category: {original_details['service_category']}",
                f"Grade: {original_details['grade']}",
                f"Years of Service: {original_details['years']}",
                f"Date Range: {original_details['start_date'].strftime('%B %d, %Y')} to {original_details['end_date'].strftime('%B %d, %Y')}",
                f"Dependent Status: {'With Dependents' if original_details['dependents'] else 'Without Dependents'}",
                f"Present for Duty: {'Yes' if present else 'No'}",
                "\nSpecial Duty Status:",
                status_text
            ]

            for info in orig_info:
                elements.append(Paragraph(info, styles['Normal']))
            elements.append(Spacer(1, 12))

            elements.append(Paragraph("Correct Pay Details", styles['Heading2']))

        print("Debug: Adding basic information...")
        # Basic Information
        info_style = ParagraphStyle('CustomBody', parent=styles['Normal'], fontSize=12, spaceAfter=6)

        correct_present = True  # For correct pay details
        status_text = get_special_duty_text(
            hazardous_duty and correct_present,
            hardship_duty and correct_present,
            at_border and correct_present,
            correct_present
        )

        basic_info = [
            f"Service Category: {service_category.value}",
            f"Grade: {military_grade}",
            f"Years of Service: {years_of_service}",
            f"Date Range: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
            f"Dependent Status: {'With Dependents' if has_dependents else 'Without Dependents'}",
            "\nSpecial Duty Status:",
            status_text
        ]

        for info in basic_info:
            elements.append(Paragraph(info, info_style))
        elements.append(Spacer(1, 20))

        print("Debug: Processing pay comparison...")
        if is_correction:
            # Display comparison
            elements.append(Paragraph("Pay Comparison", styles['Heading2']))
            
            # Calculate the difference between correct and original pay
            correct_pay = st.session_state.correct_pay['grand_total']
            original_pay = pay_info['grand_total']
            difference = correct_pay - original_pay
            
            diff_text = f"{format_currency(abs(difference))} ({'(+)' if difference > 0 else '(-)'})"
            compare_data = [
                ["Component", "Original", "Correct", "Difference"],
                ["Total Pay", 
                 format_currency(original_pay),
                 format_currency(correct_pay),
                 diff_text]
            ]

            table = Table(compare_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))

            print("Debug: Processing monthly breakdown...")
            # Monthly breakdown comparison
            elements.append(Paragraph("Monthly Breakdown Comparison", styles['Heading2']))

            # In the correction report, original pay info is in pay_info (passed parameter)
            # and correct pay info is in st.session_state.correct_pay
            # So we need to restructure how we access the monthly breakdown
            if 'monthly_breakdown' in pay_info:
                original_monthly_breakdown = pay_info['monthly_breakdown']
                correct_monthly_breakdown = st.session_state.correct_pay['monthly_breakdown']
                
                # Get all unique months from both breakdowns
                all_months = set(original_monthly_breakdown.keys()).union(
                    set(correct_monthly_breakdown.keys())
                )
                
                for month in sorted(all_months):
                    elements.append(Paragraph(f"\n{month}", styles['Heading3']))
                    
                    orig_month = original_monthly_breakdown.get(month, {})
                    corr_month = correct_monthly_breakdown.get(month, {})

                    month_data = [["Component", "Original", "Correct", "Difference"]]

                    if service_category == ServiceCategory.TEXAS_SG:
                        components = [
                            ("Base Pay", 'base_pay'),
                            ("Special Pay", 'special_pay'),
                            ("Allowances", 'allowances'),
                            ("Monthly Total", 'total')
                        ]
                    else:
                        components = [
                            ("Base Pay", 'base_pay'),
                            ("BAH", 'bah'),
                            ("BAS", 'bas'),
                            ("Per Diem", 'per_diem')  # Add Per Diem component
                        ]

                        # Add special duty components only if they exist
                        if any(key in orig_month or key in corr_month for key in ['hazard_pay', 'hardship_pay', 'danger_pay']):
                            components.extend([
                                ("Hazard Pay", 'hazard_pay'),
                                ("Hardship Pay", 'hardship_pay'),
                                ("Danger Pay", 'danger_pay')
                            ])

                        components.append(("Monthly Total", 'total'))

                    for component, key in components:
                        orig_value = orig_month.get(key, 0)
                        corr_value = corr_month.get(key, 0)
                        diff_value = corr_value - orig_value

                        month_data.append([
                            component,
                            format_currency(orig_value),
                            format_currency(corr_value),
                            format_currency(diff_value)
                        ])

                    table = Table(month_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
                    ]))
                    elements.append(table)
                    elements.append(Spacer(1, 12))

        print("Debug: Adding final summary...")
        # Grand Total
        elements.append(Spacer(1, 20))
        if is_correction:
            # Calculate the difference for the grand total section
            correct_pay = st.session_state.correct_pay['grand_total']
            original_pay = pay_info['grand_total']
            difference = correct_pay - original_pay
            
            difference_text = format_currency(abs(difference))
            status_text = "(Underpayment)" if difference > 0 else "(Overpayment)"
            total_text = f"Total Difference: {difference_text} {status_text}"
            text_color = colors.green if difference > 0 else colors.red

            elements.append(Paragraph(
                total_text,
                ParagraphStyle('Total', parent=styles['Heading2'], textColor=text_color)
            ))
        else:
            elements.append(Paragraph(
                f"Grand Total: {format_currency(pay_info['grand_total'])}",
                ParagraphStyle('Total', parent=styles['Heading2'], textColor=colors.green)
            ))

        print("Debug: Building final PDF...")
        doc.build(elements, onFirstPage=onFirstPage, onLaterPages=onLaterPages)
        print("Debug: PDF report generated successfully")
        return filename
    except Exception as e:
        print(f"Debug: Error in PDF generation - {str(e)}")
        raise

def generate_excel_report(pay_info, service_category, military_grade, years_of_service, start_date, end_date, 
                         has_dependents, hazardous_duty, hardship_duty, at_border,
                         sm_name, sm_dodid, sm_task_force, sm_company,
                         is_correction=False, original_details=None):
    """Generate an Excel report of the pay calculation"""
    filename = f"pay_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pay Calculation Report"

    # Styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    total_font = Font(bold=True)

    # Title
    ws.cell(row=1, column=1, value="SAD Pay Correction Report" if is_correction else "SAD Pay Calculator Report")
    ws.cell(row=1, column=1).font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    # Basic Information
    current_row = 3

    # Service Member Information
    if sm_name or sm_dodid or sm_task_force or sm_company:
        ws.cell(row=current_row, column=1, value="Service Member Information").font = Font(bold=True)
        current_row += 1

        if sm_name:
            ws.cell(row=current_row, column=1, value="Name:")
            ws.cell(row=current_row, column=2, value=sm_name)
            current_row += 1
        if sm_dodid:
            ws.cell(row=current_row, column=1, value="DOD ID:")
            ws.cell(row=current_row, column=2, value=sm_dodid)
            current_row += 1
        if sm_task_force:
            ws.cell(row=current_row, column=1, value="Task Force:")
            ws.cell(row=current_row, column=2, value=sm_task_force)
            current_row += 1
        if sm_company:
            ws.cell(row=current_row, column=1, value="Company:")
            ws.cell(row=current_row, column=2, value=sm_company)
            current_row += 2

    if is_correction:
        # Add original details
        ws.cell(row=current_row, column=1, value="Original Pay Details").font = Font(bold=True)
        current_row += 1

        present = original_details.get('present_this_month', False)
        orig_info = [
            ["Service Category:", original_details['service_category']],
            ["Grade:", original_details['grade']],
            ["Years of Service:", str(original_details['years'])],
            ["Date Range:", f"{original_details['start_date'].strftime('%B %d, %Y')} to {original_details['end_date'].strftime('%B %d, %Y')}"],
            ["Dependent Status:", 'With Dependents' if original_details['dependents'] else 'Without Dependents'],
            ["Present for Duty:", 'Yes' if present else 'No'],
            ["Hazardous Duty:", 'Yes' if original_details['hazardous_duty'] and present else 'No'],
            ["Hardship Duty:", 'Yes' if original_details['hardship_duty'] and present else 'No'],
            ["Imminent Danger:", 'Yes' if original_details['at_border'] and present else 'No']
        ]

        for info in orig_info:
            ws.cell(row=current_row, column=1, value=info[0])
            ws.cell(row=current_row, column=2, value=info[1])
            current_row += 1

        current_row += 1
        ws.cell(row=current_row, column=1, value="Correct Pay Details").font = Font(bold=True)
        current_row += 1

    basic_info = [
        ["Service Category:", service_category.value],
        ["Grade:", military_grade],
        ["Years of Service:", str(years_of_service)],
        ["Date Range:", f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"],
        ["Dependent Status:", 'With Dependents' if has_dependents else 'Without Dependents'],
        ["Hazardous Duty:", 'Yes' if hazardous_duty else 'No'],
        ["Hardship Duty:", 'Yes' if hardship_duty else 'No'],
        ["Imminent Danger:", 'Yes' if at_border else 'No']
    ]

    for info in basic_info:
        ws.cell(row=current_row, column=1, value=info[0])
        ws.cell(row=current_row, column=2, value=info[1])
        current_row += 1

    if is_correction:
        # Add pay comparison section
        current_row += 2
        ws.cell(row=current_row, column=1, value="Pay Comparison").font = Font(bold=True)
        current_row += 1

        # Headers
        headers = ["Component", "Original", "Correct", "Difference"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        # Total comparison
        correct_pay = st.session_state.correct_pay['grand_total']
        original_pay = pay_info['grand_total']
        difference = correct_pay - original_pay
        
        ws.cell(row=current_row, column=1, value="Total Pay")
        ws.cell(row=current_row, column=2, value=format_currency(original_pay))
        ws.cell(row=current_row, column=3, value=format_currency(correct_pay))
        ws.cell(row=current_row, column=4, value=format_currency(abs(difference)) +
                (" (+)" if difference > 0 else " (-)"))

        # Monthly breakdown comparison
        current_row += 2
        ws.cell(row=current_row, column=1, value="Monthly Breakdown Comparison").font = Font(bold=True)
        current_row += 1

        # Similar approach to PDF generation - restructuring how we access monthly breakdown
        original_monthly_breakdown = pay_info['monthly_breakdown']
        correct_monthly_breakdown = st.session_state.correct_pay['monthly_breakdown']
        
        # Get all unique months from both breakdowns
        all_months = set(original_monthly_breakdown.keys()).union(
            set(correct_monthly_breakdown.keys())
        )
        
        for month in sorted(all_months):
            orig_month = original_monthly_breakdown.get(month, {})
            corr_month = correct_monthly_breakdown.get(month, {})

            ws.cell(row=current_row, column=1, value=month).font = Font(bold=True)
            current_row += 1

            # Headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            current_row += 1

            # Data rows
            if service_category == ServiceCategory.TEXAS_SG:
                components = [
                    ("Base Pay", 'base_pay'),
                    ("Special Pay", 'special_pay'),
                    ("Allowances", 'allowances'),
                    ("Monthly Total", 'total')
                ]
            else:
                components = [
                    ("Base Pay", 'base_pay'),
                    ("BAH", 'bah'),
                    ("BAS", 'bas'),
                    ("Per Diem", 'per_diem') #Add Per Diem component
                ]

                # Add special duty components only if they exist
                if any(key in orig_month or key in corr_month for key in ['hazard_pay', 'hardship_pay', 'danger_pay']):
                    components.extend([
                        ("Hazard Pay", 'hazard_pay'),
                        ("Hardship Pay", 'hardship_pay'),
                        ("Danger Pay", 'danger_pay')
                    ])

                components.append(("Monthly Total", 'total'))

            for label, key in components:
                orig_value = orig_month.get(key, 0)
                corr_value = corr_month.get(key, 0)
                diff_value = corr_value - orig_value

                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=2, value=format_currency(orig_value))
                ws.cell(row=current_row, column=3, value=format_currency(corr_value))
                ws.cell(row=current_row, column=4, value=format_currency(diff_value))

                if label == "Monthly Total":
                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).font = total_font

                current_row += 1

            current_row += 1

    # Grand Total
    current_row += 1
    if is_correction:
        # Calculate the final difference for Grand Total
        correct_pay = st.session_state.correct_pay['grand_total']
        original_pay = pay_info['grand_total']
        difference = correct_pay - original_pay
        
        difference_text = format_currency(abs(difference))
        status_text = "(Underpayment)" if difference > 0 else "(Overpayment)"
        ws.cell(row=current_row, column=1, value="Total Difference:")
        ws.cell(row=current_row, column=2, value=f"{difference_text} {status_text}")
    else:
        ws.cell(row=current_row, column=1, value=f"Grand Total for {pay_info['total_days']} days:")
        ws.cell(row=current_row, column=2, value=format_currency(pay_info['grand_total']))

    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)

    # Add logo to excel
    try:
        img = PILImage.open('attached_assets/NEW LOGO SMALL.png')
        # Resize image for Excel
        desired_width = 200
        ratio = desired_width / float(img.size[0])
        desired_height = int(float(img.size[1]) * ratio)
        img = img.resize((desired_width, desired_height), PILImage.Resampling.LANCZOS)

        # Save resized image
        img_path = 'attached_assets/excel_logo_temp.png'
        img.save(img_path)

        # Add to Excel
        logo = XLImage(img_path)
        ws.add_image(logo, 'E3')
    except Exception as e:
        print(f"Could not add logo to Excel: {str(e)}")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width

    # Add "NOT FOR OFFICIAL USE" watermark as large, semi-transparent text
    # Using a single large cell with formatted text as the watermark
    watermark_row = 15  # Position in the middle of the sheet
    watermark_col = 3   # Position in the middle columns
    ws.cell(row=watermark_row, column=watermark_col).value = "NOT FOR OFFICIAL USE"
    ws.cell(row=watermark_row, column=watermark_col).font = Font(name='Calibri', bold=True, color='D3D3D3', size=36)  # Light gray color
    ws.cell(row=watermark_row, column=watermark_col).alignment = Alignment(horizontal='center', vertical='center', textRotation=45)  # Diagonal text

    wb.save(filename)
    return filename